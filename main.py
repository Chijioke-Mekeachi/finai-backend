import base64
import io
import json
import logging
import os
import re
import secrets
import uuid
import wave
from datetime import date, datetime, timedelta
from typing import Any

import httpx
from fastapi import Depends, FastAPI, File, Header, HTTPException, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware

# Load backend/.env automatically for local runs.
try:
    from dotenv import load_dotenv  # type: ignore

    load_dotenv(os.path.join(os.path.dirname(__file__), ".env"), override=True)
except Exception:
    pass

try:
    from .deps import CurrentUser, get_current_user
    from .schemas import (
        AdminTransactionRow,
        AdminUserDetail,
        AdminUserSummary,
        AiAnalyzeRequest,
        AiAnalyzeResponse,
        AiChatSummaryOut,
        AiMessageOut,
        AiTtsRequest,
        AiTtsResponse,
        AiVisionRequest,
        PaystackInitializeRequest,
        PaystackInitializeResponse,
        PaystackVerifyResponse,
    )
    from .supabase_rest import (
        filter_eq,
        filter_gte,
        filter_in,
        filter_like,
        filter_lte,
        filter_neq,
        rest_count,
        rest_delete,
        rest_insert,
        rest_rpc,
        rest_select,
        rest_update,
    )
except ImportError:
    from deps import CurrentUser, get_current_user
    from schemas import (
        AdminTransactionRow,
        AdminUserDetail,
        AdminUserSummary,
        AiAnalyzeRequest,
        AiAnalyzeResponse,
        AiChatSummaryOut,
        AiMessageOut,
        AiTtsRequest,
        AiTtsResponse,
        AiVisionRequest,
        PaystackInitializeRequest,
        PaystackInitializeResponse,
        PaystackVerifyResponse,
    )
    from supabase_rest import (
        filter_eq,
        filter_gte,
        filter_in,
        filter_like,
        filter_lte,
        filter_neq,
        rest_count,
        rest_delete,
        rest_insert,
        rest_rpc,
        rest_select,
        rest_update,
    )

app = FastAPI(title="FinTrack Pro API", version="2.0.0")

origins = [
    o.strip()
    for o in os.getenv("CORS_ORIGINS", "http://localhost:3000,http://127.0.0.1:3000").split(",")
    if o.strip()
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

logger = logging.getLogger("fintrack")

_AI_CHAT_LIMITS_24H: dict[str, int] = {
    "basic": 10,
    "standard": 30,
    "premium": 80,
    "strategic": 30,
    "executive": 80,
}

_SUBSCRIPTION_PLAN_PRICES: dict[str, float] = {
    "standard": 12.0,
    "strategic": 25.0,
    "executive": 75.0,
}


def _ai_chat_limit_for_plan(plan_id: str | None) -> tuple[str, int]:
    plan = (plan_id or "").strip().lower() or "basic"
    if plan == "strategic":
        plan = "standard"
    elif plan == "executive":
        plan = "premium"
    return plan, _AI_CHAT_LIMITS_24H.get(plan, _AI_CHAT_LIMITS_24H["basic"])


def _parse_plan_amounts(value: str | None) -> dict[str, int]:
    raw = (value or "").strip()
    if not raw:
        return {}
    out: dict[str, int] = {}
    for part in raw.split(","):
        p = part.strip()
        if not p or "=" not in p:
            continue
        k, v = p.split("=", 1)
        key = k.strip().lower()
        try:
            amt = int(v.strip())
        except Exception:
            continue
        if key and amt > 0:
            out[key] = amt
    return out


def _paystack_secret() -> str:
    secret = (os.getenv("PAYSTACK_SECRET_KEY") or "").strip()
    if not secret:
        raise HTTPException(status_code=503, detail="Paystack not configured")
    return secret


def _paystack_currency() -> str:
    return (os.getenv("PAYSTACK_CURRENCY") or "USD").strip().upper() or "USD"


def _paystack_amount_for_plan(plan_id: str) -> int:
    plan = (plan_id or "").strip().lower()
    if plan not in _SUBSCRIPTION_PLAN_PRICES:
        raise HTTPException(status_code=422, detail="Invalid plan_id")

    overrides = _parse_plan_amounts(os.getenv("PAYSTACK_PLAN_AMOUNTS"))
    if plan in overrides:
        return overrides[plan]

    price = _SUBSCRIPTION_PLAN_PRICES[plan]
    return int(round(price * 100))


def _truncate_err(value: Any, max_len: int = 240) -> str:
    try:
        s = str(value)
    except Exception:
        return "error"
    s = " ".join(s.split())
    return s if len(s) <= max_len else (s[:max_len] + "…")


def _normalize_model(model: str) -> str:
    model = (model or "").strip()
    if model.startswith("models/"):
        return model[len("models/") :]
    return model


def _get_model(requested: str | None) -> str:
    return _normalize_model(
        requested
        or os.getenv("GEMINI_MODEL", "").strip()
        or "gemini-2.0-flash"
    )


def _get_gemini_api_version() -> str:
    return (os.getenv("GEMINI_API_VERSION", "v1beta") or "v1beta").strip()


def _get_tts_model(requested: str | None) -> str:
    return _normalize_model(
        requested
        or os.getenv("GEMINI_TTS_MODEL", "").strip()
        or "gemini-2.5-flash-preview-tts"
    )


def _get_tts_voice(requested: str | None) -> str:
    return (requested or os.getenv("GEMINI_TTS_VOICE", "").strip() or "Kore").strip()


def _parse_sample_rate_from_mime(mime_type: str | None, fallback: int = 24000) -> int:
    mime = (mime_type or "").strip().lower()
    if not mime:
        return fallback
    m = re.search(r"rate=(\d+)", mime)
    if not m:
        return fallback
    try:
        rate = int(m.group(1))
        return rate if 8000 <= rate <= 48000 else fallback
    except Exception:
        return fallback


def _pcm_s16le_to_wav_base64(pcm_bytes: bytes, sample_rate_hz: int) -> str:
    buf = io.BytesIO()
    with wave.open(buf, "wb") as wf:
        wf.setnchannels(1)
        wf.setsampwidth(2)
        wf.setframerate(sample_rate_hz)
        wf.writeframes(pcm_bytes)
    return base64.b64encode(buf.getvalue()).decode("ascii")


def _ai_error_detail(response: httpx.Response) -> str:
    debug = os.getenv("DEBUG_AI_ERRORS", "0") == "1"
    if not debug:
        return "AI service error"
    try:
        data = response.json()
        err = data.get("error") if isinstance(data, dict) else None
        if isinstance(err, dict):
            msg = err.get("message") or "AI service error"
            status_txt = err.get("status")
            return f"{status_txt}: {msg}" if status_txt else str(msg)
    except Exception:
        pass
    text = (response.text or "").strip().replace("\n", " ")
    if len(text) > 300:
        text = text[:300] + "…"
    return f"AI service error ({response.status_code}): {text}" if text else f"AI service error ({response.status_code})"


def _extract_json_object(text: str) -> dict[str, Any] | None:
    raw = (text or "").strip()
    if not raw:
        return None
    if raw.startswith("```"):
        lines = raw.splitlines()
        if lines:
            lines = lines[1:]
        if lines and lines[-1].strip().startswith("```"):
            lines = lines[:-1]
        raw = "\n".join(lines).strip()

    start = raw.find("{")
    end = raw.rfind("}")
    if start == -1 or end == -1 or end <= start:
        return None
    candidate = raw[start : end + 1]
    try:
        parsed = json.loads(candidate)
        return parsed if isinstance(parsed, dict) else None
    except Exception:
        return None


def _assistant_text_for_storage(text: str) -> str:
    extracted = _extract_json_object(text)
    if not extracted:
        return text
    if isinstance(extracted.get("assistant_markdown"), str) and extracted["assistant_markdown"].strip():
        return extracted["assistant_markdown"].strip()
    if isinstance(extracted.get("message"), str) and extracted["message"].strip():
        return extracted["message"].strip()
    return text


def require_admin(x_admin_key: str | None = Header(default=None, alias="X-Admin-Key")) -> None:
    expected = (os.getenv("ADMIN_API_KEY") or "").strip()
    if not expected:
        raise HTTPException(status_code=503, detail="Admin access not configured")
    if not x_admin_key or x_admin_key.strip() != expected:
        raise HTTPException(status_code=401, detail="Invalid admin key")


@app.get("/health")
def health() -> dict[str, str]:
    return {"Status": "Good"}


@app.get("/api/me")
def api_me(current_user: CurrentUser = Depends(get_current_user)) -> dict[str, Any]:
    return {
        "id": current_user.id,
        "email": current_user.email,
        "name": current_user.name,
        "subscription_plan_id": current_user.subscription_plan_id,
    }


@app.get("/admin/ai/models")
async def admin_list_ai_models(
    version: str | None = None,
    _: None = Depends(require_admin),
) -> dict[str, Any]:
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=503, detail="AI service not configured")

    api_version = (version or _get_gemini_api_version()).strip()
    url = f"https://generativelanguage.googleapis.com/{api_version}/models"
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            res = await client.get(url, params={"key": api_key})
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail=f"AI request failed: {exc}") from exc

    if res.status_code >= 400:
        raise HTTPException(status_code=502, detail=_ai_error_detail(res))

    payload = res.json()
    models = payload.get("models", []) if isinstance(payload, dict) else []
    supported: list[dict[str, Any]] = []
    for m in models:
        if not isinstance(m, dict):
            continue
        methods = m.get("supportedGenerationMethods") or []
        if "generateContent" not in methods:
            continue
        supported.append(
            {
                "name": _normalize_model(m.get("name", "")),
                "displayName": m.get("displayName"),
                "methods": methods,
            }
        )
    suggestion = next((m["name"] for m in supported if "flash" in (m["name"] or "").lower()), None)
    return {
        "api_version": api_version,
        "env_model": os.getenv("GEMINI_MODEL"),
        "suggested_model": suggestion,
        "models": supported,
    }


@app.get("/admin/users", response_model=list[AdminUserSummary])
async def admin_list_users(
    limit: int = 200,
    offset: int = 0,
    _: None = Depends(require_admin),
) -> list[AdminUserSummary]:
    limit = min(max(limit, 1), 500)
    offset = max(offset, 0)
    rows = await rest_rpc(fn="admin_list_users", payload={"p_limit": limit, "p_offset": offset}, auth="admin")
    out: list[AdminUserSummary] = []
    for r in rows or []:
        out.append(
            AdminUserSummary(
                id=str(r.get("id")),
                email=str(r.get("email") or "unknown@example.com"),
                name=str(r.get("name") or "Member"),
                subscription_plan_id=str(r.get("subscription_plan_id") or "standard"),
                created_at=r.get("created_at"),
                transactions_count=int(r.get("transactions_count") or 0),
                settings=(
                    {
                        "company_name": r.get("company_name"),
                        "currency": r.get("currency"),
                        "fiscal_year_start": r.get("fiscal_year_start"),
                        "tax_rate": float(r.get("tax_rate") or 0),
                    }
                    if r.get("company_name") is not None
                    else None
                ),
            )
        )
    return out


@app.get("/admin/users/{user_id}", response_model=AdminUserDetail)
async def admin_get_user(
    user_id: str,
    limit: int = 200,
    offset: int = 0,
    _: None = Depends(require_admin),
) -> AdminUserDetail:
    profile = await rest_select(
        table="profiles",
        select="id,email,full_name,subscription_plan_id,created_at",
        filters=[filter_eq("id", user_id)],
        single=True,
        auth="admin",
    )
    if not profile:
        raise HTTPException(status_code=404, detail="User not found")

    settings = await rest_select(
        table="business_settings",
        select="company_name,currency,fiscal_year_start,tax_rate",
        filters=[filter_eq("user_id", user_id)],
        single=True,
        auth="admin",
    )

    tx_rows = await rest_select(
        table="transactions",
        select="id,date,type,category,amount,entity,description,created_at",
        filters=[filter_eq("user_id", user_id)],
        order="date.desc,created_at.desc",
        limit=min(max(limit, 1), 1000),
        offset=max(offset, 0),
        auth="admin",
    )

    return AdminUserDetail(
        id=str(profile.get("id")),
        email=str(profile.get("email") or "unknown@example.com"),
        name=str(profile.get("full_name") or "Member"),
        subscription_plan_id=str(profile.get("subscription_plan_id") or "standard"),
        created_at=profile.get("created_at"),
        transactions_count=int((await rest_count(table="transactions", filters=[filter_eq("user_id", user_id)], auth="admin")).exact),
        settings=(
            {
                "company_name": settings.get("company_name"),
                "currency": settings.get("currency"),
                "fiscal_year_start": settings.get("fiscal_year_start"),
                "tax_rate": float(settings.get("tax_rate") or 0),
            }
            if isinstance(settings, dict) and settings
            else None
        ),
        transactions=[
            {
                "id": t.get("id"),
                "date": t.get("date"),
                "type": t.get("type"),
                "category": t.get("category"),
                "amount": float(t.get("amount") or 0),
                "entity": t.get("entity"),
                "description": t.get("description"),
            }
            for t in (tx_rows or [])
            if isinstance(t, dict)
        ],
    )


@app.get("/admin/transactions", response_model=list[AdminTransactionRow])
async def admin_list_transactions(
    from_date: date | None = None,
    to_date: date | None = None,
    type: str | None = None,
    q: str | None = None,
    limit: int = 500,
    offset: int = 0,
    format: str | None = None,
    _: None = Depends(require_admin),
):
    limit = min(max(limit, 1), 5000)
    offset = max(offset, 0)

    filters: list[tuple[str, str]] = []
    if from_date is not None:
        filters.append(filter_gte("date", from_date.isoformat()))
    if to_date is not None:
        filters.append(filter_lte("date", to_date.isoformat()))
    if type is not None and type.strip():
        filters.append(filter_eq("type", type.strip()))

    rows = await rest_select(
        table="transactions",
        select="id,user_id,date,type,category,amount,entity,description,created_at",
        filters=filters,
        order="date.desc,created_at.desc",
        limit=limit,
        offset=offset,
        auth="admin",
    )

    tx_rows = [r for r in (rows or []) if isinstance(r, dict)]
    user_ids = sorted({str(r.get("user_id")) for r in tx_rows if r.get("user_id")})
    profiles_by_id: dict[str, dict[str, Any]] = {}
    if user_ids:
        profs = await rest_select(
            table="profiles",
            select="id,email,full_name,subscription_plan_id",
            filters=[filter_in("id", user_ids)],
            auth="admin",
        )
        for p in profs or []:
            if isinstance(p, dict) and p.get("id"):
                profiles_by_id[str(p["id"])] = p

    needle = (q or "").strip().lower()
    out: list[AdminTransactionRow] = []
    for tx in tx_rows:
        uid = str(tx.get("user_id") or "")
        prof = profiles_by_id.get(uid, {})
        row = AdminTransactionRow(
            id=str(tx.get("id")),
            user_id=uid,
            user_email=str(prof.get("email") or "unknown@example.com"),
            user_name=str(prof.get("full_name") or "Member"),
            user_plan_id=str(prof.get("subscription_plan_id") or "standard"),
            date=tx.get("date"),
            type=tx.get("type"),
            category=tx.get("category"),
            amount=float(tx.get("amount") or 0),
            entity=tx.get("entity"),
            description=tx.get("description"),
            created_at=tx.get("created_at"),
        )

        if needle:
            hay = " ".join(
                [
                    str(row.user_email or "").lower(),
                    str(row.user_name or "").lower(),
                    str(row.entity or "").lower(),
                    str(row.category or "").lower(),
                ]
            )
            if needle not in hay:
                continue

        out.append(row)

    if (format or "").lower() == "csv":
        import csv

        out_buf = io.StringIO()
        writer = csv.writer(out_buf)
        writer.writerow(
            [
                "id",
                "user_id",
                "user_email",
                "user_name",
                "user_plan_id",
                "date",
                "type",
                "category",
                "amount",
                "entity",
                "description",
                "created_at",
            ]
        )
        for r in out:
            writer.writerow(
                [
                    r.id,
                    r.user_id,
                    r.user_email,
                    r.user_name,
                    r.user_plan_id,
                    r.date,
                    r.type,
                    r.category,
                    f"{r.amount:.2f}",
                    r.entity,
                    r.description or "",
                    r.created_at,
                ]
            )
        from fastapi.responses import Response

        filename = f"fintrack_admin_transactions_{datetime.utcnow().date().isoformat()}.csv"
        return Response(
            content=out_buf.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
        )

    return out


async def _enforce_ai_chat_limit(
    *,
    current_user: CurrentUser,
    endpoint: str,
    model: str | None,
    purpose: str | None,
) -> str:
    now = datetime.utcnow()
    window_start = now - timedelta(hours=24)

    canonical_plan, limit = _ai_chat_limit_for_plan(current_user.subscription_plan_id)

    used = (
        await rest_count(
            table="ai_api_calls",
            filters=[
                filter_eq("user_id", current_user.id),
                filter_eq("category", "chat"),
                filter_neq("status", "failed"),
                filter_gte("created_at", window_start.isoformat() + "Z"),
            ],
            auth="admin",
        )
    ).exact

    if used >= limit:
        oldest = await rest_select(
            table="ai_api_calls",
            select="created_at",
            filters=[
                filter_eq("user_id", current_user.id),
                filter_eq("category", "chat"),
                filter_neq("status", "failed"),
                filter_gte("created_at", window_start.isoformat() + "Z"),
            ],
            order="created_at.asc",
            limit=1,
            auth="admin",
        )
        reset_at = now + timedelta(hours=24)
        if isinstance(oldest, list) and oldest and isinstance(oldest[0], dict) and oldest[0].get("created_at"):
            try:
                ts = str(oldest[0]["created_at"]).replace("Z", "+00:00")
                reset_at = datetime.fromisoformat(ts) + timedelta(hours=24)
            except Exception:
                pass

        retry_after = max(1, int((reset_at - now).total_seconds()))
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail={
                "message": "AI chat call limit exceeded",
                "limit_24h": limit,
                "used_24h": used,
                "plan_id": current_user.subscription_plan_id,
                "tier": canonical_plan,
            },
            headers={"Retry-After": str(retry_after)},
        )

    call_id = str(uuid.uuid4())
    await rest_insert(
        table="ai_api_calls",
        rows=[
            {
                "id": call_id,
                "user_id": current_user.id,
                "category": "chat",
                "endpoint": endpoint,
                "model": model,
                "purpose": purpose,
                "status": "started",
                "created_at": now.isoformat() + "Z",
                "updated_at": now.isoformat() + "Z",
            }
        ],
        auth="admin",
    )
    return call_id


async def _update_ai_call(call_id: str, *, status_value: str, error: str | None = None) -> None:
    patch: dict[str, Any] = {"status": status_value, "updated_at": datetime.utcnow().isoformat() + "Z"}
    if error:
        patch["error"] = error
    await rest_update(table="ai_api_calls", patch=patch, filters=[filter_eq("id", call_id)], auth="admin")


@app.post("/api/ai/analyze", response_model=AiAnalyzeResponse)
async def ai_analyze(
    payload: AiAnalyzeRequest,
    current_user: CurrentUser = Depends(get_current_user),
) -> AiAnalyzeResponse:
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=503, detail="AI service not configured")

    model = _get_model(payload.model)
    api_version = _get_gemini_api_version()
    url = f"https://generativelanguage.googleapis.com/{api_version}/models/{model}:generateContent"
    purpose = (payload.purpose or "general").strip() or "general"

    call_id = await _enforce_ai_chat_limit(
        current_user=current_user,
        endpoint="/api/ai/analyze",
        model=model,
        purpose=purpose,
    )

    try:
        await rest_insert(
            table="ai_messages",
            rows=[
                {
                    "user_id": current_user.id,
                    "purpose": purpose,
                    "role": "user",
                    "content": payload.user_message or payload.prompt,
                }
            ],
            auth="admin",
        )

        body = {"contents": [{"parts": [{"text": payload.prompt}]}]}
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                logger.info("Gemini analyze request user=%s version=%s model=%s", current_user.id, api_version, model)
                response = await client.post(url, params={"key": api_key}, json=body)
        except httpx.HTTPError as exc:
            raise HTTPException(status_code=502, detail=f"AI request failed: {exc}") from exc

        if response.status_code >= 400:
            raise HTTPException(status_code=502, detail=_ai_error_detail(response))

        data = response.json()
        candidates = data.get("candidates", [])
        if not candidates:
            raise HTTPException(status_code=502, detail="AI response empty")

        text = candidates[0].get("content", {}).get("parts", [{}])[0].get("text", "")
        stored_assistant = _assistant_text_for_storage(text)

        await rest_insert(
            table="ai_messages",
            rows=[
                {
                    "user_id": current_user.id,
                    "purpose": purpose,
                    "role": "assistant",
                    "content": stored_assistant,
                }
            ],
            auth="admin",
        )
        await _update_ai_call(call_id, status_value="succeeded")
        return AiAnalyzeResponse(text=text)
    except Exception as exc:
        await _update_ai_call(call_id, status_value="failed", error=_truncate_err(getattr(exc, "detail", None) or exc))
        raise


@app.post("/api/ai/vision", response_model=AiAnalyzeResponse)
async def ai_vision(
    payload: AiVisionRequest,
    current_user: CurrentUser = Depends(get_current_user),
) -> AiAnalyzeResponse:
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=503, detail="AI service not configured")

    model = _get_model(payload.model)
    api_version = _get_gemini_api_version()
    url = f"https://generativelanguage.googleapis.com/{api_version}/models/{model}:generateContent"
    purpose = (payload.purpose or "general").strip() or "general"

    call_id = await _enforce_ai_chat_limit(
        current_user=current_user,
        endpoint="/api/ai/vision",
        model=model,
        purpose=purpose,
    )

    try:
        await rest_insert(
            table="ai_messages",
            rows=[
                {
                    "user_id": current_user.id,
                    "purpose": purpose,
                    "role": "user",
                    "content": payload.user_message or f"[vision] {payload.prompt}",
                }
            ],
            auth="admin",
        )

        body = {
            "contents": [
                {
                    "parts": [
                        {"inlineData": {"mimeType": payload.mime_type, "data": payload.image_base64}},
                        {"text": payload.prompt},
                    ]
                }
            ]
        }
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                logger.info("Gemini vision request user=%s version=%s model=%s", current_user.id, api_version, model)
                response = await client.post(url, params={"key": api_key}, json=body)
        except httpx.HTTPError as exc:
            raise HTTPException(status_code=502, detail=f"AI request failed: {exc}") from exc

        if response.status_code >= 400:
            raise HTTPException(status_code=502, detail=_ai_error_detail(response))

        data = response.json()
        candidates = data.get("candidates", [])
        if not candidates:
            raise HTTPException(status_code=502, detail="AI response empty")

        text = candidates[0].get("content", {}).get("parts", [{}])[0].get("text", "")
        stored_assistant = _assistant_text_for_storage(text)
        await rest_insert(
            table="ai_messages",
            rows=[
                {
                    "user_id": current_user.id,
                    "purpose": purpose,
                    "role": "assistant",
                    "content": stored_assistant,
                }
            ],
            auth="admin",
        )
        await _update_ai_call(call_id, status_value="succeeded")
        return AiAnalyzeResponse(text=text)
    except Exception as exc:
        await _update_ai_call(call_id, status_value="failed", error=_truncate_err(getattr(exc, "detail", None) or exc))
        raise


@app.post("/api/ai/tts", response_model=AiTtsResponse)
async def ai_tts(
    payload: AiTtsRequest,
    current_user: CurrentUser = Depends(get_current_user),
) -> AiTtsResponse:
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=503, detail="AI service not configured")

    model = _get_tts_model(payload.model)
    api_version = _get_gemini_api_version()
    url = f"https://generativelanguage.googleapis.com/{api_version}/models/{model}:generateContent"

    voice_name = _get_tts_voice(payload.voice_name)
    lang = (payload.language_code or "").strip()

    tts_text = (payload.text or "").strip()
    if payload.style and payload.style.strip():
        tts_text = f"{payload.style.strip()}\n\n{tts_text}"

    generation_config: dict[str, Any] = {
        "responseModalities": ["AUDIO"],
        "speechConfig": {"voiceConfig": {"prebuiltVoiceConfig": {"voiceName": voice_name}}},
    }
    if lang:
        generation_config["speechConfig"]["languageCode"] = lang

    body = {"contents": [{"parts": [{"text": tts_text}]}], "generationConfig": generation_config}
    headers = {"x-goog-api-key": api_key}

    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            logger.info("Gemini tts request user=%s version=%s model=%s voice=%s", current_user.id, api_version, model, voice_name)
            response = await client.post(url, headers=headers, json=body)
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail=f"AI request failed: {exc}") from exc

    if response.status_code >= 400:
        raise HTTPException(status_code=502, detail=_ai_error_detail(response))

    data = response.json()
    candidates = data.get("candidates", [])
    if not candidates:
        raise HTTPException(status_code=502, detail="AI response empty")

    part0 = candidates[0].get("content", {}).get("parts", [{}])[0]
    inline = part0.get("inlineData") or part0.get("inline_data")
    if not isinstance(inline, dict):
        raise HTTPException(status_code=502, detail="AI response did not include audio data")

    audio_b64 = inline.get("data")
    mime_type = inline.get("mimeType") or inline.get("mime_type")
    if not isinstance(audio_b64, str) or not audio_b64.strip():
        raise HTTPException(status_code=502, detail="AI response did not include audio data")

    try:
        pcm_bytes = base64.b64decode(audio_b64)
    except Exception as exc:
        raise HTTPException(status_code=502, detail="AI response audio was not valid base64") from exc

    sample_rate_hz = _parse_sample_rate_from_mime(mime_type, fallback=24000)
    wav_b64 = _pcm_s16le_to_wav_base64(pcm_bytes, sample_rate_hz=sample_rate_hz)
    return AiTtsResponse(audio_base64=wav_b64, mime_type="audio/wav", sample_rate_hz=sample_rate_hz)


@app.get("/api/ai/messages", response_model=list[AiMessageOut])
async def list_ai_messages(
    purpose: str | None = None,
    limit: int = 100,
    offset: int = 0,
    current_user: CurrentUser = Depends(get_current_user),
) -> list[AiMessageOut]:
    limit = min(max(limit, 1), 500)
    offset = max(offset, 0)

    filters: list[tuple[str, str]] = [filter_eq("user_id", current_user.id)]
    if purpose and purpose.strip():
        filters.append(filter_eq("purpose", purpose.strip()))

    rows = await rest_select(
        table="ai_messages",
        select="id,role,purpose,content,created_at",
        filters=filters,
        order="created_at.desc",
        limit=limit,
        offset=offset,
        auth="admin",
    )
    rows = [r for r in (rows or []) if isinstance(r, dict)]
    rows.reverse()
    return [
        AiMessageOut(
            id=r.get("id"),
            role=r.get("role"),
            purpose=r.get("purpose"),
            content=r.get("content"),
            created_at=r.get("created_at"),
        )
        for r in rows
    ]


@app.get("/api/ai/chats", response_model=list[AiChatSummaryOut])
async def list_ai_chats(
    prefix: str = "advisor_chat",
    limit: int = 50,
    current_user: CurrentUser = Depends(get_current_user),
) -> list[AiChatSummaryOut]:
    limit = min(max(limit, 1), 200)
    prefix = (prefix or "").strip()
    rows = await rest_rpc(
        fn="list_ai_chats",
        payload={"p_prefix": prefix, "p_limit": limit},
        auth="user",
        access_token=current_user.access_token,
    )
    out: list[AiChatSummaryOut] = []
    for r in rows or []:
        if not isinstance(r, dict):
            continue
        out.append(
            AiChatSummaryOut(
                id=r.get("id"),
                purpose=r.get("purpose"),
                title=r.get("title"),
                created_at=r.get("created_at"),
                updated_at=r.get("updated_at"),
                messages_count=int(r.get("messages_count") or 0),
            )
        )
    return out


@app.delete("/api/ai/messages", status_code=204)
async def delete_ai_messages(
    purpose: str | None = None,
    current_user: CurrentUser = Depends(get_current_user),
) -> None:
    filters: list[tuple[str, str]] = [filter_eq("user_id", current_user.id)]
    if purpose and purpose.strip():
        filters.append(filter_eq("purpose", purpose.strip()))
    await rest_delete(table="ai_messages", filters=filters, auth="admin")


async def _compare_and_persist_excel_scan(
    *,
    file1: UploadFile,
    file2: UploadFile,
    current_user: CurrentUser,
) -> tuple[str, list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception:
        raise HTTPException(status_code=503, detail="Excel compare not configured (missing openpyxl)")

    max_bytes = int(os.getenv("AUDIT_MAX_UPLOAD_BYTES", str(10 * 1024 * 1024)))
    max_diffs = int(os.getenv("AUDIT_MAX_DIFFS", "5000"))
    max_cells = int(os.getenv("AUDIT_MAX_CELLS", str(2_000_000)))
    row_context_max_cols = int(os.getenv("AUDIT_ROW_CONTEXT_MAX_COLS", "200"))
    row_context_max_diffs = int(os.getenv("AUDIT_ROW_CONTEXT_MAX_DIFFS", "1000"))
    effective_max_diffs = min(max_diffs, row_context_max_diffs)

    b1 = await file1.read()
    b2 = await file2.read()
    if len(b1) > max_bytes or len(b2) > max_bytes:
        raise HTTPException(status_code=413, detail="File too large")

    scan_id = str(uuid.uuid4())
    await rest_insert(
        table="audit_scans",
        rows=[
            {
                "id": scan_id,
                "user_id": current_user.id,
                "file1_name": (file1.filename or "file1.xlsx")[:255],
                "file2_name": (file2.filename or "file2.xlsx")[:255],
                "differences_count": 0,
                "status": "started",
            }
        ],
        auth="admin",
    )

    diffs: list[dict[str, Any]] = []
    total_cells = 0

    try:
        wb1 = load_workbook(io.BytesIO(b1), data_only=False, read_only=True)
        wb2 = load_workbook(io.BytesIO(b2), data_only=False, read_only=True)
    except Exception as exc:
        await rest_update(
            table="audit_scans",
            patch={"status": "failed", "error": _truncate_err(exc)},
            filters=[filter_eq("id", scan_id)],
            auth="admin",
        )
        raise HTTPException(status_code=400, detail=f"Invalid Excel file(s): {exc}") from exc

    def excel_cell_value_to_text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        return str(value)

    def format_excel_diff_line(
        *,
        sheet_name: str,
        col_letter: str,
        row: int,
        col_header: str | None,
        row_header: str | None,
        value_text: str,
    ) -> str:
        parts = []
        if sheet_name:
            parts.append(sheet_name)
        parts.append(f"{col_letter}, {row}")
        header_bits: list[str] = []
        if col_header and col_header.strip():
            header_bits.append(f"col:{col_header.strip()}")
        if row_header and row_header.strip():
            header_bits.append(f"row:{row_header.strip()}")
        if header_bits:
            parts.append(f"[{' | '.join(header_bits)}]")
        return f"{' '.join(parts)} = {value_text}"

    try:
        sheets = sorted(set(wb1.sheetnames) | set(wb2.sheetnames))
        batch: list[dict[str, Any]] = []
        for sheet_name in sheets:
            ws1 = wb1[sheet_name] if sheet_name in wb1.sheetnames else None
            ws2 = wb2[sheet_name] if sheet_name in wb2.sheetnames else None

            max_row = max(getattr(ws1, "max_row", 0) or 0, getattr(ws2, "max_row", 0) or 0)
            max_col = max(getattr(ws1, "max_column", 0) or 0, getattr(ws2, "max_column", 0) or 0)

            total_cells += max_row * max_col
            if total_cells > max_cells:
                raise HTTPException(status_code=413, detail="Excel file too large to compare (cell limit exceeded)")

            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    v1 = ws1.cell(row=row, column=col).value if ws1 else None
                    v2 = ws2.cell(row=row, column=col).value if ws2 else None
                    if v1 == v2:
                        continue

                    t1 = excel_cell_value_to_text(v1)
                    t2 = excel_cell_value_to_text(v2)
                    if not t1 and not t2:
                        continue

                    col_header_val = None
                    row_header_val = None
                    if row != 1:
                        h1 = ws1.cell(row=1, column=col).value if ws1 else None
                        h2 = ws2.cell(row=1, column=col).value if ws2 else None
                        col_header_val = excel_cell_value_to_text(h1) or excel_cell_value_to_text(h2) or None
                    if col != 1:
                        rh1 = ws1.cell(row=row, column=1).value if ws1 else None
                        rh2 = ws2.cell(row=row, column=1).value if ws2 else None
                        row_header_val = excel_cell_value_to_text(rh1) or excel_cell_value_to_text(rh2) or None

                    col_letter = get_column_letter(col)

                    def build_row_context(ws: Any | None) -> tuple[list[dict[str, Any]], bool]:
                        if not ws:
                            return [], False
                        cols = int(getattr(ws, "max_column", 0) or 0)
                        cols = max(cols, max_col)
                        if cols <= 0:
                            return [], False
                        capped = min(cols, row_context_max_cols)
                        truncated = cols > capped
                        out_cells: list[dict[str, Any]] = []
                        for c in range(1, capped + 1):
                            header_val = excel_cell_value_to_text(ws.cell(row=1, column=c).value)
                            header = header_val.strip() if header_val.strip() else get_column_letter(c)
                            cell_val = excel_cell_value_to_text(ws.cell(row=row, column=c).value)
                            out_cells.append(
                                {
                                    "col": c,
                                    "col_letter": get_column_letter(c),
                                    "header": header,
                                    "value": cell_val,
                                }
                            )
                        return out_cells, truncated

                    file1_row, file1_row_truncated = build_row_context(ws1)
                    file2_row, file2_row_truncated = build_row_context(ws2)
                    row_truncated = file1_row_truncated or file2_row_truncated

                    line1 = format_excel_diff_line(
                        sheet_name=sheet_name,
                        col_letter=col_letter,
                        row=row,
                        col_header=col_header_val,
                        row_header=row_header_val,
                        value_text=t1,
                    )
                    line2 = format_excel_diff_line(
                        sheet_name=sheet_name,
                        col_letter=col_letter,
                        row=row,
                        col_header=col_header_val,
                        row_header=row_header_val,
                        value_text=t2,
                    )

                    diffs.append(
                        {
                            "file1": line1,
                            "file2": line2,
                            "sheet": sheet_name,
                            "row": row,
                            "col": col,
                            "col_letter": col_letter,
                            "col_header": col_header_val,
                            "row_header": row_header_val,
                            "file1_value": t1,
                            "file2_value": t2,
                            "file1_row": file1_row,
                            "file2_row": file2_row,
                            "row_truncated": row_truncated,
                        }
                    )

                    batch.append(
                        {
                            "scan_id": scan_id,
                            "sheet_name": (sheet_name or "")[:120],
                            "row": row,
                            "col": col,
                            "col_header": col_header_val,
                            "row_header": row_header_val,
                            "file1_value": t1,
                            "file2_value": t2,
                            "file1_row_json": file1_row,
                            "file2_row_json": file2_row,
                            "row_truncated": row_truncated,
                        }
                    )

                    if len(batch) >= 250:
                        await rest_insert(table="audit_scan_diffs", rows=batch, auth="admin")
                        batch = []

                    if len(diffs) >= effective_max_diffs:
                        raise HTTPException(status_code=413, detail=f"Too many differences (limit {effective_max_diffs})")

        if batch:
            await rest_insert(table="audit_scan_diffs", rows=batch, auth="admin")

        await rest_update(
            table="audit_scans",
            patch={"differences_count": len(diffs), "status": "succeeded"},
            filters=[filter_eq("id", scan_id)],
            auth="admin",
        )
    except Exception as exc:
        await rest_update(
            table="audit_scans",
            patch={"status": "failed", "error": _truncate_err(getattr(exc, "detail", None) or exc)},
            filters=[filter_eq("id", scan_id)],
            auth="admin",
        )
        raise
    finally:
        try:
            wb1.close()
        except Exception:
            pass
        try:
            wb2.close()
        except Exception:
            pass

    logger.info("audit_compare_excel user=%s scan_id=%s diffs=%s", current_user.id, scan_id, len(diffs))
    return scan_id, diffs


@app.post("/api/audit/compare-excel")
async def audit_compare_excel(
    file1: UploadFile = File(...),
    file2: UploadFile = File(...),
    current_user: CurrentUser = Depends(get_current_user),
) -> list[dict[str, str]]:
    _, diffs = await _compare_and_persist_excel_scan(file1=file1, file2=file2, current_user=current_user)
    return [{"file1": str(d.get("file1", "")), "file2": str(d.get("file2", ""))} for d in diffs]


@app.post("/api/audit/compare-excel/v2")
async def audit_compare_excel_v2(
    file1: UploadFile = File(...),
    file2: UploadFile = File(...),
    current_user: CurrentUser = Depends(get_current_user),
) -> dict[str, Any]:
    scan_id, diffs = await _compare_and_persist_excel_scan(file1=file1, file2=file2, current_user=current_user)
    return {"scan_id": scan_id, "diffs": diffs}


@app.get("/api/audit/scans")
async def list_audit_scans(
    limit: int = 25,
    offset: int = 0,
    current_user: CurrentUser = Depends(get_current_user),
) -> list[dict[str, Any]]:
    limit = min(max(limit, 1), 200)
    offset = max(offset, 0)
    rows = await rest_select(
        table="audit_scans",
        select="id,file1_name,file2_name,differences_count,status,error,created_at",
        filters=[filter_eq("user_id", current_user.id)],
        order="created_at.desc",
        limit=limit,
        offset=offset,
        auth="admin",
    )
    return [
        {
            "id": r.get("id"),
            "file1_name": r.get("file1_name"),
            "file2_name": r.get("file2_name"),
            "differences_count": int(r.get("differences_count") or 0),
            "status": r.get("status"),
            "error": r.get("error"),
            "created_at": r.get("created_at"),
        }
        for r in (rows or [])
        if isinstance(r, dict)
    ]


@app.get("/api/audit/scans/{scan_id}")
async def get_audit_scan(
    scan_id: str,
    current_user: CurrentUser = Depends(get_current_user),
) -> dict[str, Any]:
    scan = await rest_select(
        table="audit_scans",
        select="id,file1_name,file2_name,differences_count,status,error,created_at",
        filters=[filter_eq("id", scan_id), filter_eq("user_id", current_user.id)],
        single=True,
        auth="admin",
    )
    if not scan:
        raise HTTPException(status_code=404, detail="Scan not found")

    diffs = await rest_select(
        table="audit_scan_diffs",
        select="sheet_name,row,col,col_header,row_header,file1_value,file2_value,file1_row_json,file2_row_json,row_truncated,created_at",
        filters=[filter_eq("scan_id", scan_id)],
        order="created_at.asc",
        auth="admin",
    )

    try:
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception:
        def get_column_letter(n: int) -> str:  # type: ignore
            return str(n)

    out = []
    for d in diffs or []:
        if not isinstance(d, dict):
            continue
        col_letter = get_column_letter(int(d.get("col") or 0))
        out.append(
            {
                "sheet": d.get("sheet_name"),
                "row": int(d.get("row") or 0),
                "col": int(d.get("col") or 0),
                "col_letter": col_letter,
                "col_header": d.get("col_header"),
                "row_header": d.get("row_header"),
                "file1_value": d.get("file1_value") or "",
                "file2_value": d.get("file2_value") or "",
                "file1_row": d.get("file1_row_json") if isinstance(d.get("file1_row_json"), list) else [],
                "file2_row": d.get("file2_row_json") if isinstance(d.get("file2_row_json"), list) else [],
                "row_truncated": bool(d.get("row_truncated") or False),
            }
        )

    return {
        "id": scan.get("id"),
        "file1_name": scan.get("file1_name"),
        "file2_name": scan.get("file2_name"),
        "differences_count": int(scan.get("differences_count") or 0),
        "status": scan.get("status"),
        "error": scan.get("error"),
        "created_at": scan.get("created_at"),
        "diffs": out,
    }


@app.post("/api/billing/paystack/initialize", response_model=PaystackInitializeResponse)
async def paystack_initialize(
    payload: PaystackInitializeRequest,
    current_user: CurrentUser = Depends(get_current_user),
) -> PaystackInitializeResponse:
    secret = _paystack_secret()
    if not (current_user.email or "").strip():
        raise HTTPException(status_code=422, detail="Email missing for billing")

    plan_id = (payload.plan_id or "").strip().lower()
    amount = _paystack_amount_for_plan(plan_id)
    currency = _paystack_currency()
    reference = secrets.token_hex(16)
    callback_url = (payload.callback_url or "").strip() or None

    await rest_insert(
        table="payment_transactions",
        rows=[
            {
                "user_id": current_user.id,
                "provider": "paystack",
                "reference": reference,
                "plan_id": plan_id,
                "amount": amount,
                "currency": currency,
                "status": "started",
                "raw": None,
            }
        ],
        auth="admin",
    )

    body: dict[str, Any] = {
        "email": current_user.email,
        "amount": amount,
        "currency": currency,
        "reference": reference,
        "metadata": {"user_id": current_user.id, "plan_id": plan_id},
    }
    if callback_url:
        body["callback_url"] = callback_url

    headers = {"Authorization": f"Bearer {secret}", "Content-Type": "application/json"}
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post("https://api.paystack.co/transaction/initialize", json=body, headers=headers)
    except httpx.HTTPError as exc:
        await rest_update(
            table="payment_transactions",
            patch={"status": "failed", "raw": {"error": str(exc)}},
            filters=[filter_eq("reference", reference)],
            auth="admin",
        )
        raise HTTPException(status_code=502, detail=f"Paystack request failed: {exc}") from exc

    raw_json: Any
    try:
        raw_json = resp.json()
    except Exception:
        raw_json = {"status_code": resp.status_code, "text": (resp.text or "")[:2000]}

    await rest_update(
        table="payment_transactions",
        patch={"raw": raw_json if isinstance(raw_json, dict) else {"raw": raw_json}},
        filters=[filter_eq("reference", reference)],
        auth="admin",
    )

    if resp.status_code >= 400:
        await rest_update(
            table="payment_transactions",
            patch={"status": "failed"},
            filters=[filter_eq("reference", reference)],
            auth="admin",
        )
        raise HTTPException(status_code=502, detail="Paystack initialize failed")

    data = raw_json.get("data") if isinstance(raw_json, dict) else None
    if not isinstance(data, dict):
        raise HTTPException(status_code=502, detail="Paystack initialize returned invalid response")

    authorization_url = data.get("authorization_url")
    access_code = data.get("access_code")
    returned_ref = data.get("reference") or reference
    if not isinstance(authorization_url, str) or not isinstance(access_code, str) or not isinstance(returned_ref, str):
        raise HTTPException(status_code=502, detail="Paystack initialize returned invalid response")

    return PaystackInitializeResponse(
        authorization_url=authorization_url,
        access_code=access_code,
        reference=returned_ref,
    )


@app.get("/api/billing/paystack/verify", response_model=PaystackVerifyResponse)
async def paystack_verify(
    reference: str,
    current_user: CurrentUser = Depends(get_current_user),
) -> PaystackVerifyResponse:
    ref = (reference or "").strip()
    if not ref:
        raise HTTPException(status_code=422, detail="reference is required")

    tx = await rest_select(
        table="payment_transactions",
        select="reference,plan_id,amount,currency,status,raw",
        filters=[
            filter_eq("reference", ref),
            filter_eq("user_id", current_user.id),
            filter_eq("provider", "paystack"),
        ],
        single=True,
        auth="admin",
    )
    if not tx:
        raise HTTPException(status_code=404, detail="Payment reference not found")

    if tx.get("status") == "succeeded":
        return PaystackVerifyResponse(status="succeeded", plan_id=tx.get("plan_id"), reference=ref)

    secret = _paystack_secret()
    headers = {"Authorization": f"Bearer {secret}"}
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.get(f"https://api.paystack.co/transaction/verify/{ref}", headers=headers)
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail=f"Paystack request failed: {exc}") from exc

    raw_json: Any
    try:
        raw_json = resp.json()
    except Exception:
        raw_json = {"status_code": resp.status_code, "text": (resp.text or "")[:2000]}

    await rest_update(
        table="payment_transactions",
        patch={"raw": raw_json if isinstance(raw_json, dict) else {"raw": raw_json}},
        filters=[filter_eq("reference", ref)],
        auth="admin",
    )

    if resp.status_code >= 400:
        await rest_update(
            table="payment_transactions",
            patch={"status": "failed"},
            filters=[filter_eq("reference", ref)],
            auth="admin",
        )
        raise HTTPException(status_code=502, detail="Paystack verify failed")

    data = raw_json.get("data") if isinstance(raw_json, dict) else None
    if not isinstance(data, dict):
        raise HTTPException(status_code=502, detail="Paystack verify returned invalid response")

    paid_ok = (data.get("status") == "success")
    paid_amount = data.get("amount")
    paid_currency = data.get("currency")
    paid_email = (data.get("customer") or {}).get("email") if isinstance(data.get("customer"), dict) else None
    metadata = data.get("metadata") if isinstance(data.get("metadata"), dict) else {}

    if not paid_ok:
        await rest_update(table="payment_transactions", patch={"status": "failed"}, filters=[filter_eq("reference", ref)], auth="admin")
        raise HTTPException(status_code=400, detail="Payment not successful")

    if isinstance(paid_amount, int) and int(tx.get("amount") or 0) != paid_amount:
        await rest_update(table="payment_transactions", patch={"status": "failed"}, filters=[filter_eq("reference", ref)], auth="admin")
        raise HTTPException(status_code=400, detail="Payment amount mismatch")

    if isinstance(paid_currency, str) and paid_currency.strip().upper() != str(tx.get("currency") or "").upper():
        await rest_update(table="payment_transactions", patch={"status": "failed"}, filters=[filter_eq("reference", ref)], auth="admin")
        raise HTTPException(status_code=400, detail="Payment currency mismatch")

    if isinstance(paid_email, str) and current_user.email:
        if paid_email.strip().lower() != current_user.email.strip().lower():
            await rest_update(table="payment_transactions", patch={"status": "failed"}, filters=[filter_eq("reference", ref)], auth="admin")
            raise HTTPException(status_code=400, detail="Payment customer mismatch")

    meta_user_id = metadata.get("user_id") if isinstance(metadata, dict) else None
    if isinstance(meta_user_id, str) and meta_user_id != current_user.id:
        await rest_update(table="payment_transactions", patch={"status": "failed"}, filters=[filter_eq("reference", ref)], auth="admin")
        raise HTTPException(status_code=400, detail="Payment metadata mismatch")

    plan_id = str(tx.get("plan_id") or "").strip().lower()
    if not plan_id:
        raise HTTPException(status_code=502, detail="Payment row missing plan_id")

    await rest_update(
        table="profiles",
        patch={"subscription_plan_id": plan_id, "updated_at": datetime.utcnow().isoformat() + "Z"},
        filters=[filter_eq("id", current_user.id)],
        auth="admin",
    )
    await rest_update(
        table="payment_transactions",
        patch={"status": "succeeded"},
        filters=[filter_eq("reference", ref)],
        auth="admin",
    )
    return PaystackVerifyResponse(status="succeeded", plan_id=plan_id, reference=ref)

